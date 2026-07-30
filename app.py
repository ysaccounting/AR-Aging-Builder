"""
AR Aging Report — Flask backend.

Upload an invoices export (.xlsx). Select a group and an As of Date.
Downloads a workbook with two tabs — AR Aging Summary and Invoice Details.
"""

import io
import os
import re
import time
import uuid
import shutil
import tempfile
import datetime as dt

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, jsonify, send_file, send_from_directory, abort

app = Flask(__name__, static_folder=None)
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STORE_DIR  = os.path.join(tempfile.gettempdir(), "ar_aging_store")
os.makedirs(STORE_DIR, exist_ok=True)

# ── Constants ─────────────────────────────────────────────────────────────────

CLIENT_NAME_MAP = {
    "SeatGeek":        "SeatGeek",
    "Gametime":        "Gametime",
    "GoTickets":       "GoTickets",
    "Mercury":         "Mercury",
    "StubHub":         "Stubhub",
    "Ticket Evolution":"Ticket Evolution",
    "TicketNetwork":   "TicketNetwork",
    "TicketsNow":      "TicketsNow",
    "TickPick":        "TickPick",
    "Vivid Seats":     "Vivid Seats",
}

NETWORK_ORDER = [
    "Gametime", "GoTickets", "Mercury", "Offsite", "SeatGeek",
    "Stubhub", "Ticket Evolution", "TicketNetwork", "TicketsNow",
    "TickPick", "Vivid Seats",
]

BUCKETS = ["Current", "1 to 30", "31 to 60", "61 to 90", "91 and Over"]

COMPANY_RENAMES = {
    "YS Tickets Spec": "YS Tickets",
    "YSA 2":           "YSA",
    "YSA 3":           "YSA",
}

GROUP_OPTIONS = ["Y&S Group", "YS-SeatGeek", "The Ticket Guy"]

# ── File store cleanup ────────────────────────────────────────────────────────

def _cleanup_old(max_age_seconds=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        path = os.path.join(STORE_DIR, name)
        try:
            if os.path.isdir(path) and now - os.path.getmtime(path) > max_age_seconds:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            pass


def _safe_filename(s):
    return re.sub(r'[\\/:*?"<>|]+', " ", s).strip() if s else s


# ── Data processing ───────────────────────────────────────────────────────────

def assign_bucket(days: int) -> str:
    if days <= 0:   return "Current"
    elif days <= 30: return "1 to 30"
    elif days <= 60: return "31 to 60"
    elif days <= 90: return "61 to 90"
    else:            return "91 and Over"


def detect_format(df: pd.DataFrame) -> str:
    """'full' = original export with Paid/IsCancelled; 'light' = pre-filtered 8-col export."""
    return "full" if ("Paid" in df.columns and "IsCancelled" in df.columns) else "light"


def load_and_filter(data: bytes, as_of_date: pd.Timestamp) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(data))

    # Normalize Created to datetime regardless of format (string, datetime, etc.)
    df["Created"] = pd.to_datetime(df["Created"], errors="coerce")

    fmt = detect_format(df)

    if fmt == "full":
        unpaid = df[
            (df["Paid"] == "No") &
            (df["IsCancelled"] == "No") &
            (df["Bal."] > 0)
        ].copy()
    else:
        unpaid = df[df["Bal."] > 0].copy()

    # Exclude balances under $1
    unpaid = unpaid[unpaid["Bal."] >= 1].copy()

    # Remove duplicates: same Client + Ext Order # (only when Ext Order # is not blank)
    has_ext = unpaid["Ext Order #"].notna() & (unpaid["Ext Order #"].astype(str).str.strip() != "")
    with_ext    = unpaid[has_ext].drop_duplicates(subset=["Client", "Ext Order #"], keep="first")
    without_ext = unpaid[~has_ext]
    unpaid = pd.concat([with_ext, without_ext]).sort_index()

    unpaid["days_out"]         = (as_of_date - unpaid["Created"]).dt.days
    unpaid["bucket"]           = unpaid["days_out"].apply(assign_bucket)
    unpaid["is_other_network"] = ~unpaid["Client"].isin(CLIENT_NAME_MAP)
    return unpaid


def build_pivot(unpaid: pd.DataFrame):
    main = unpaid[~unpaid["is_other_network"]].copy()
    main["display_name"] = main["Client"].map(CLIENT_NAME_MAP)
    pivot = main.pivot_table(
        index="display_name", columns="bucket", values="Bal.",
        aggfunc="sum", fill_value=0,
    )
    other_by_bucket = unpaid[unpaid["is_other_network"]].groupby("bucket")["Bal."].sum()
    return pivot, other_by_bucket


def get_val(row_name, bkt, pivot, other_by_bucket) -> float:
    if row_name == "Offsite":
        return other_by_bucket.get(bkt, 0.0)
    if row_name in pivot.index and bkt in pivot.columns:
        return pivot.loc[row_name, bkt]
    return 0.0


# ── Excel builder ─────────────────────────────────────────────────────────────

def _styles():
    thick = Side(style="medium", color="000000")
    return {
        "title":        Font(name="Arial", bold=True, size=14),
        "subtitle":     Font(name="Arial", bold=True, size=12),
        "date":         Font(name="Arial", size=11),
        "header":       Font(name="Arial", bold=True, size=11),
        "body":         Font(name="Arial", size=11),
        "total":        Font(name="Arial", bold=True, size=11),
        "center":       Alignment(horizontal="center", vertical="center"),
        "header_border":Border(top=thick, bottom=thick),
        "total_border": Border(top=thick, bottom=thick),
    }


def _build_summary_sheet(wb, as_of_date, active_rows, group_name):
    ws = wb.active
    ws.title = "AR Aging Summary"
    s = _styles()

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    for r, (text, font) in enumerate([
        (group_name,                                    s["title"]),
        ("A/R Aging Summary",                           s["subtitle"]),
        (f'As of {as_of_date.strftime("%B %-d, %Y")}', s["date"]),
    ], 1):
        ws.merge_cells(f"A{r}:G{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font = font
        c.alignment = s["center"]

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8

    for col_idx, h in enumerate(["Network","Current","1 to 30","31 to 60","61 to 90","91 and Over","Total"], 1):
        c = ws.cell(row=5, column=col_idx, value=h)
        c.font = s["header"]
        c.alignment = s["center"]
        c.border = s["header_border"]
    ws.row_dimensions[5].height = 18

    # SUMIFS formulas referencing Invoice Details tab
    detail   = "'Invoice Details'"
    net_col  = f"{detail}!$C:$C"
    amt_col  = f"{detail}!$E:$E"
    age_col  = f"{detail}!$H:$H"

    for i, row_name in enumerate(active_rows):
        r = 6 + i
        ws.row_dimensions[r].height = 16
        c = ws.cell(row=r, column=1, value=row_name)
        c.font = s["body"]
        c.alignment = s["center"]

        for col_idx, bkt in enumerate(BUCKETS, 2):
            c = ws.cell(row=r, column=col_idx,
                        value=f'=SUMIFS({amt_col},{net_col},"{row_name}",{age_col},"{bkt}")')
            c.font = s["body"]
            c.alignment = s["center"]
            c.number_format = "$#,##0"

        c = ws.cell(row=r, column=7, value=f"=SUM(B{r}:{get_column_letter(2+len(BUCKETS)-1)}{r})")
        c.font = s["body"]
        c.alignment = s["center"]
        c.number_format = "$#,##0"

    total_row = 6 + len(active_rows)
    ws.row_dimensions[total_row].height = 18
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = s["total"]
    c.alignment = s["center"]
    c.border = s["total_border"]

    for col_idx in range(2, 8):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}6:{col_letter}{total_row-1})")
        c.font = s["total"]
        c.alignment = s["center"]
        c.border = s["total_border"]
        c.number_format = "$#,##0"


def _build_detail_sheet(wb, unpaid: pd.DataFrame):
    ws = wb.create_sheet("Invoice Details")
    s = _styles()

    source_df = unpaid.copy()
    source_df["Company"]       = source_df["Company"].replace(COMPANY_RENAMES)
    source_df["Client_display"]= source_df["Client"].apply(lambda c: CLIENT_NAME_MAP.get(c, "Offsite"))

    src_cols = ["Company", "Inv#", "Client_display", "Ext Order #", "Bal.", "Status", "Created", "bucket"]
    headers  = {
        "Company":        "Broker",
        "Inv#":           "Invoice #",
        "Client_display": "Network",
        "Ext Order #":    "Ext Order #",
        "Bal.":           "Amount",
        "Status":         "Status",
        "Created":        "Invoice Date",
        "bucket":         "Aging",
    }
    widths = {"Broker":22,"Invoice #":14,"Network":18,"Ext Order #":20,
              "Amount":16,"Status":12,"Invoice Date":20,"Aging":14}

    for col_idx, col in enumerate(src_cols, 1):
        label = headers[col]
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = s["header"]
        c.alignment = s["center"]
        c.border = s["header_border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[label]

    for row_idx, row in enumerate(source_df[src_cols].itertuples(index=False), 2):
        for col_idx, (col, val) in enumerate(zip(src_cols, row), 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = s["body"]
            c.alignment = s["center"]
            if col == "Bal.":
                c.number_format = "$#,##0.00"
            elif col == "Created":
                c.number_format = "MM/DD/YYYY"

    ws.freeze_panes = "A2"


def build_workbook(data: bytes, as_of_date: pd.Timestamp, group_name: str):
    unpaid = load_and_filter(data, as_of_date)
    pivot, other_by_bucket = build_pivot(unpaid)

    active_rows = [
        rn for rn in NETWORK_ORDER
        if sum(get_val(rn, bkt, pivot, other_by_bucket) for bkt in BUCKETS) > 0
    ]

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, as_of_date, active_rows, group_name)
    _build_detail_sheet(wb, unpaid)

    buf = io.BytesIO()
    wb.save(buf)

    grand_total = sum(
        get_val(rn, bkt, pivot, other_by_bucket)
        for rn in active_rows for bkt in BUCKETS
    )
    return buf.getvalue(), len(unpaid), round(grand_total, 2)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/process", methods=["POST"])
def process():
    f = request.files.get("invoices")
    if not f or not f.filename:
        return jsonify({"error": "Please upload an invoices file."}), 400

    group_name = (request.form.get("group_name") or "").strip() or "AR Aging"

    raw_date = (request.form.get("as_of_date") or "").strip()
    if raw_date:
        try:
            as_of_date = pd.Timestamp(dt.datetime.strptime(raw_date, "%Y-%m-%d").date())
        except ValueError:
            return jsonify({"error": "As of Date must be YYYY-MM-DD."}), 400
    else:
        as_of_date = pd.Timestamp(dt.date.today())

    try:
        data = f.read()
        excel_bytes, row_count, grand_total = build_workbook(data, as_of_date, group_name)
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    token     = uuid.uuid4().hex
    safe_name = _safe_filename(group_name)
    filename  = f"AR Aging - {safe_name} - {as_of_date.strftime('%m-%d-%Y')}.xlsx"
    folder    = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)

    with open(os.path.join(folder, filename), "wb") as fh:
        fh.write(excel_bytes)

    _cleanup_old()

    return jsonify({
        "filename":     filename,
        "download_url": f"/download/{token}",
        "row_count":    row_count,
        "grand_total":  grand_total,
    })


@app.route("/download/<token>")
def download(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    files = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    if not files:
        abort(404)
    pick = files[0]
    return send_file(
        os.path.join(folder, pick),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=pick,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
